"""
Forensic validation 01 — source file structure & column mapping.

Read-only. Does NOT import modules.kpis. Reads the original workbook and
reports the physical column layout of every sheet.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

SRC = Path(r"C:\Users\PC\Downloads\Business Case.xlsx")
MODEL = Path(__file__).resolve().parent.parent / "powerbi" / "DiDi_CX_PowerBI_Model.xlsx"


def hr(title: str) -> None:
    print("\n" + "=" * 100)
    print(title)
    print("=" * 100)


def main() -> None:
    print(f"Source: {SRC}  exists={SRC.exists()}")
    print(f"Model : {MODEL}  exists={MODEL.exists()}")

    xl = pd.ExcelFile(SRC)
    hr("SHEETS IN SOURCE WORKBOOK")
    for s in xl.sheet_names:
        df = pd.read_excel(SRC, sheet_name=s)
        print(f"  {s!r:20s} rows={len(df):>7,}  cols={len(df.columns)}")

    qa = pd.read_excel(SRC, sheet_name="QA")

    hr("QA TAB — FULL POSITIONAL COLUMN MAP (Excel letter = 1-based position)")
    for i, c in enumerate(qa.columns):
        excel_idx = i + 1
        letter = get_column_letter(excel_idx)
        dtype = str(qa[c].dtype)
        nulls = int(qa[c].isna().sum())
        uniq = qa[c].dropna().unique()
        sample = sorted(uniq.tolist())[:6] if len(uniq) <= 40 else f"<{len(uniq)} distinct>"
        print(f"  {letter:>3s} (idx {excel_idx:>2d})  {str(c)[:58]:58s} {dtype:>10s} nulls={nulls:>4d}  {sample}")

    # ---- positional ranges from the business case -------------------------
    hr("POSITIONAL RANGES PER BUSINESS CASE")
    cols = list(qa.columns)

    def rng(a: int, b: int) -> list[str]:
        return [str(c) for c in cols[a - 1 : b]]

    phone_pos = rng(23, 34)   # W..AH
    chat_pos = rng(35, 42)    # AI..AP

    print(f"\nW(23)..AH(34)  -> {len(phone_pos)} columns (PHONE per business case):")
    for i, c in enumerate(phone_pos):
        print(f"   {get_column_letter(23 + i):>3s}  {c}")

    print(f"\nAI(35)..AP(42) -> {len(chat_pos)} columns (LIVE CHAT per business case):")
    for i, c in enumerate(chat_pos):
        print(f"   {get_column_letter(35 + i):>3s}  {c}")

    # ---- compare against config -------------------------------------------
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from config import LIVECHAT_ATTRS, PHONE_ATTRS  # noqa: E402

    hr("CONFIG vs POSITIONAL — PHONE")
    print(f"config PHONE_ATTRS n={len(PHONE_ATTRS)} | positional W-AH n={len(phone_pos)}")
    print(f"SAME SET?    {set(PHONE_ATTRS) == set(phone_pos)}")
    print(f"SAME ORDER?  {list(PHONE_ATTRS) == list(phone_pos)}")
    print(f"in config not in W-AH: {sorted(set(PHONE_ATTRS) - set(phone_pos))}")
    print(f"in W-AH not in config: {sorted(set(phone_pos) - set(PHONE_ATTRS))}")

    hr("CONFIG vs POSITIONAL — LIVE CHAT")
    print(f"config LIVECHAT_ATTRS n={len(LIVECHAT_ATTRS)} | positional AI-AP n={len(chat_pos)}")
    print(f"SAME SET?    {set(LIVECHAT_ATTRS) == set(chat_pos)}")
    print(f"SAME ORDER?  {list(LIVECHAT_ATTRS) == list(chat_pos)}")
    print(f"in config not in AI-AP: {sorted(set(LIVECHAT_ATTRS) - set(chat_pos))}")
    print(f"in AI-AP not in config: {sorted(set(chat_pos) - set(LIVECHAT_ATTRS))}")

    hr("OVERLAP CHECK")
    print(f"Phone ∩ LiveChat (config): {sorted(set(PHONE_ATTRS) & set(LIVECHAT_ATTRS))}")
    print(f"Phone ∩ LiveChat (positional): {sorted(set(phone_pos) & set(chat_pos))}")

    # ---- critical detection -----------------------------------------------
    hr("CRITICAL ATTRIBUTE DETECTION — every column in the whole QA tab containing 'critical'")
    crit_all = [(get_column_letter(i + 1), str(c)) for i, c in enumerate(cols) if "critical" in str(c).lower()]
    for letter, c in crit_all:
        in_phone = c in phone_pos
        in_chat = c in chat_pos
        where = "PHONE(W-AH)" if in_phone else ("LIVECHAT(AI-AP)" if in_chat else "OUTSIDE ATTR RANGE !!!")
        print(f"  {letter:>3s}  {c:70s} {where}")
    print(f"\nTotal columns with 'critical' in name: {len(crit_all)}")

    crit_in_ranges = [c for c in phone_pos + chat_pos if "critical" in c.lower()]
    print(f"Of those, inside the two attribute ranges: {len(crit_in_ranges)}")
    outside = [c for _, c in crit_all if c not in phone_pos and c not in chat_pos]
    print(f"'Critical' columns OUTSIDE attribute ranges (false positives risk): {outside}")

    # non-critical attr columns
    noncrit = [c for c in phone_pos + chat_pos if "critical" not in c.lower()]
    print(f"\nNon-critical attribute columns ({len(noncrit)}):")
    for c in noncrit:
        print(f"   {c}")

    # ---- dim_attribute in the exported model -------------------------------
    if MODEL.exists():
        hr("EXPORTED MODEL — dim_attribute")
        da = pd.read_excel(MODEL, sheet_name="dim_attribute")
        print(da.to_string(index=False))
        print(f"\nRows: {len(da)}")
        model_phone = da[da["Channel_Scope"] == "Phone"]["Attribute_Key"].tolist()
        model_chat = da[da["Channel_Scope"] == "Live Chat"]["Attribute_Key"].tolist()
        print(f"\nmodel Phone == positional W-AH ? {set(model_phone) == set(phone_pos)}")
        print(f"model LiveChat == positional AI-AP ? {set(model_chat) == set(chat_pos)}")
        mism = da[(da["Is_Critical"] == 1) != (da["Attribute_Key"].str.lower().str.contains("critical"))]
        print(f"\nRows where Is_Critical flag disagrees with the name rule: {len(mism)}")
        if len(mism):
            print(mism.to_string(index=False))

    # ---- CSAT / RECONTACT column maps --------------------------------------
    for sheet in ("CSAT", "Recontact"):
        df = pd.read_excel(SRC, sheet_name=sheet)
        hr(f"{sheet} TAB — POSITIONAL COLUMN MAP  (rows={len(df):,})")
        for i, c in enumerate(df.columns):
            letter = get_column_letter(i + 1)
            uniq = df[c].dropna().unique()
            sample = sorted(map(str, uniq.tolist()))[:5] if len(uniq) <= 25 else f"<{len(uniq)} distinct>"
            print(f"  {letter:>3s} (idx {i+1:>2d})  {str(c)[:50]:50s} {str(df[c].dtype):>10s} "
                  f"nulls={int(df[c].isna().sum()):>5d}  {sample}")


if __name__ == "__main__":
    main()
